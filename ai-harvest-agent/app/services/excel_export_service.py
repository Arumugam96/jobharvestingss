"""
Excel Export Service — generates a multi-sheet XLSX workbook from harvest results.

Workbook structure
──────────────────
Sheet 1 — Combined Jobs   (all sources, all fields + lead intelligence columns)
Sheet 2 — LinkedIn Jobs
Sheet 3 — Naukri Jobs
Sheet 4 — Dice Jobs
Sheet 5 — Lead Intelligence   (all jobs with lead status column)

Output path:  data/results/excel/YYYYMMDD_HHMMSS_harvest.xlsx
"""
from __future__ import annotations

import re
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

import structlog

# openpyxl rejects XML control characters — strip them from every cell value
_ILLEGAL_XML = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f￾￿]')

logger = structlog.get_logger(__name__)

_EXCEL_DIR = Path("data/results/excel")

# ── Column definitions ────────────────────────────────────────────────────────

_JOB_COLUMNS: list[tuple[str, str]] = [
    ("job_title",              "Job Title"),
    ("company",                "Company"),
    ("company_url",            "Company URL"),
    ("location",               "Location"),
    ("salary",                 "Salary"),
    ("experience",             "Experience"),
    ("posted_date",            "Posted Date"),
    ("job_url",                "Job URL"),
    ("job_description",        "Job Description"),
    ("skills",                 "Skills"),
    ("work_mode",              "Work Mode"),
    ("source",                 "Source"),
    ("employment_type",        "Employment Type"),
    ("job_type",               "Job Type"),
    ("domain",                 "Domain"),
    ("hiring_entity",          "Hiring Entity"),
    ("is_gcc",                 "Is GCC"),
    ("verification_status",    "Verification Status"),
    ("job_poster_name",        "Job Poster Name"),
    ("job_poster_designation", "Job Poster Designation"),
    ("linkedin_profile_url",   "LinkedIn Profile URL"),
    ("current_company",        "Current Company"),
    ("email_id",               "Email ID"),
    ("contact_number",         "Contact Number"),
]

_LEAD_COLUMNS: list[tuple[str, str]] = [
    ("job_title",              "Job Title"),
    ("company",                "Company"),
    ("source",                 "Source"),
    ("job_poster_name",        "Job Poster Name"),
    ("job_poster_designation", "Job Poster Designation"),
    ("linkedin_profile_url",   "LinkedIn Profile URL"),
    ("current_company",        "Current Company"),
    ("email_id",               "Email ID"),
    ("contact_number",         "Contact Number"),
    ("hiring_entity",          "Hiring Entity"),
    ("verification_status",    "Verification Status"),
    ("job_url",                "Job URL"),
    ("posted_date",            "Posted Date"),
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _sanitize(val: Any) -> Any:
    """Strip illegal XML control characters so openpyxl never raises IllegalCharacterError."""
    if isinstance(val, str):
        val = _ILLEGAL_XML.sub("", val)
        # Truncate very long strings (job descriptions) to 5000 chars
        if len(val) > 5000:
            val = val[:5000] + "…"
    return val


def _field(job: Any, key: str) -> Any:
    """Read a field off either a UnifiedJob-style object or a plain dict —
    the DB-backed report path passes scraped_job_view() dicts instead of
    dataclass instances."""
    if isinstance(job, dict):
        return job.get(key)
    return getattr(job, key, None)


def _lead_status(job: Any) -> str:
    """Compute Lead Status for a job record."""
    has_name  = bool(_field(job, "job_poster_name"))
    has_email = bool(_field(job, "email_id"))
    has_phone = bool(_field(job, "contact_number"))
    has_url   = bool(_field(job, "linkedin_profile_url"))
    if has_email or has_phone:
        return "Enriched - Contact Available"
    if has_name and (has_url or _field(job, "current_company")):
        return "Enriched - Profile Only"
    if has_name:
        return "Partial - Name Only"
    return "Pending"


def _job_to_row(job: Any, columns: list[tuple[str, str]]) -> list[Any]:
    """Convert a UnifiedJob (or job dict) to a flat list aligned with `columns`."""
    row: list[Any] = []
    for field_key, _ in columns:
        if field_key == "_lead_status":
            val = _lead_status(job)
        else:
            val = _field(job, field_key)
            if isinstance(val, list):
                val = ", ".join(str(v) for v in val)
            elif val is None:
                val = ""
        row.append(_sanitize(val))
    return row


def _apply_header_style(ws: Any, header_fill: str = "1F4E79") -> None:
    """Bold white text on dark-blue background for the header row."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        fill = PatternFill("solid", fgColor=header_fill)
        font = Font(bold=True, color="FFFFFF")
        align = Alignment(wrap_text=True, vertical="center")
        for cell in ws[1]:
            cell.fill  = fill
            cell.font  = font
            cell.alignment = align
    except Exception:
        pass  # openpyxl not available — skip styling


def _autofit_columns(ws: Any, max_width: int = 60) -> None:
    """Set reasonable column widths based on content."""
    try:
        from openpyxl.utils import get_column_letter
        for col_idx, col_cells in enumerate(ws.columns, 1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                (len(str(cell.value or "")) for cell in col_cells),
                default=10,
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, max_width)
    except Exception:
        pass


def _write_sheet(ws: Any, columns: list[tuple[str, str]], jobs: list[Any]) -> None:
    """Write header + data rows to a worksheet."""
    ws.append([display for _, display in columns])
    for job in jobs:
        ws.append(_job_to_row(job, columns))
        logger.debug("excel_row_written", sheet=ws.title, rows=ws.max_row - 1)
    _apply_header_style(ws)
    _autofit_columns(ws)
    ws.freeze_panes = "A2"
    logger.info("sheet_completed", sheet=ws.title, rows=ws.max_row - 1)


# ══════════════════════════════════════════════════════════════════════════════
# ExcelExportService
# ══════════════════════════════════════════════════════════════════════════════

class ExcelExportService:
    """Generate a multi-sheet Excel workbook from harvest results."""

    def _build_workbook(
        self,
        all_jobs:       list[Any],
        jobs_by_source: dict[str, list[Any]],
    ) -> "Any":
        """Assemble the 5-sheet workbook. Jobs may be UnifiedJob instances or
        scraped_job_view() dicts (the DB-backed report path)."""
        import openpyxl

        wb = openpyxl.Workbook()

        # ── Sheet 1: Combined Jobs ─────────────────────────────────────────────
        ws_combined = wb.active
        ws_combined.title = "Combined Jobs"
        _write_sheet(ws_combined, _JOB_COLUMNS, all_jobs)

        # ── Sheets 2-4: per-source ─────────────────────────────────────────────
        for sheet_title in ("LinkedIn Jobs", "Naukri Jobs", "Dice Jobs"):
            source_key = sheet_title.replace(" Jobs", "")  # "LinkedIn" | "Naukri" | "Dice"
            ws = wb.create_sheet(title=sheet_title)
            source_jobs = jobs_by_source.get(source_key, [])
            if source_key == "LinkedIn":
                logger.info("linkedin_sheet_created", sheet=sheet_title, job_count=len(source_jobs))
            _write_sheet(ws, _JOB_COLUMNS, source_jobs)
            if source_key == "LinkedIn":
                logger.info("linkedin_jobs_written_to_excel", rows=len(source_jobs))

        # ── Sheet 5: Lead Intelligence (ALL jobs with lead status column) ────
        ws_leads = wb.create_sheet(title="Lead Intelligence")
        _write_sheet(ws_leads, _LEAD_COLUMNS, all_jobs)
        return wb

    def build_bytes(
        self,
        all_jobs:       list[Any],
        jobs_by_source: dict[str, list[Any]] | None = None,
    ) -> bytes:
        """Build the workbook fully in memory and return the .xlsx bytes —
        used by GET /download/excel and the report email, which generate the
        report from the DB on demand instead of a harvest-time file."""
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            logger.error("excel_export_skipped", reason="openpyxl not installed — run: pip install openpyxl")
            return b""

        if jobs_by_source is None:
            jobs_by_source = {}
            for j in all_jobs:
                jobs_by_source.setdefault(_field(j, "source") or "", []).append(j)

        wb  = self._build_workbook(all_jobs, jobs_by_source)
        buf = BytesIO()
        wb.save(buf)
        logger.info("excel_built_in_memory", total_jobs=len(all_jobs), bytes=buf.getbuffer().nbytes)
        return buf.getvalue()

    def export(
        self,
        all_jobs:       list[Any],
        jobs_by_source: dict[str, list[Any]],
        run_id:         str,
        filters_snap:   dict,
    ) -> str:
        """
        Build the workbook and write it to data/results/excel/.
        Returns the absolute path of the saved file.
        """
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            logger.error("excel_export_skipped", reason="openpyxl not installed — run: pip install openpyxl")
            return ""

        _EXCEL_DIR.mkdir(parents=True, exist_ok=True)
        ts       = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"{ts}_harvest.xlsx"
        path     = _EXCEL_DIR / filename

        wb = self._build_workbook(all_jobs, jobs_by_source)

        enriched_count = sum(
            1 for j in all_jobs
            if _field(j, "job_poster_name")
            or _field(j, "email_id")
            or _field(j, "contact_number")
        )

        wb.save(str(path))
        logger.info(
            "excel_exported",
            path         = str(path.resolve()),
            total_jobs   = len(all_jobs),
            lead_records = enriched_count,
            sheets       = ["Combined Jobs", "LinkedIn Jobs", "Naukri Jobs", "Dice Jobs", "Lead Intelligence"],
        )
        logger.info("harvest_completed", total_jobs=len(all_jobs), lead_records=enriched_count, excel_generated=True)
        return str(path.resolve())

    def export_path_for_run(self, run_id: str) -> str:
        """Return the expected Excel path for a run_id (may not exist yet)."""
        _EXCEL_DIR.mkdir(parents=True, exist_ok=True)
        return str((_EXCEL_DIR / f"{run_id}_harvest.xlsx").resolve())
